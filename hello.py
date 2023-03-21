from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#wb = Workbook()

#load existing sppredsheet
wb = load_workbook('hello.xlsx')
#createworksheet object
ws = wb.active
#print something from our spreadsheet
name = ws['A2'].value
color = ws['B2'].value
#print(f'{name}: {color}')

#grab a whole column value with for loop
column_a = ws['1'] #get the list of column A tuple
#for cell in column_a:
#    print(f'{cell.value}\n')

# Grab a range 
range = ws['A2':'A10']
for cell in range:
    for x in cell:
        print(x.value)